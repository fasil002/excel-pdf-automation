import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

folder_path = input("Enter the folder path that contains messy text files: ")
all_files = glob.glob(os.path.join(folder_path, "*.csv"))

if not all_files:
    print("No CSV/text files found in that folder.")
else:
    df_list = []
    for file in all_files:
        try:
            # read messy files separated by spaces/tabs
            df = pd.read_csv(file, sep=r"\s+", engine="python")

            # remove completely empty rows
            df = df.dropna(how="all")
            df = df[~df.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)]

            df_list.append(df)
        except Exception as e:
            print(f"Error reading {file}: {e}")

    combined = pd.concat(df_list, ignore_index=True)

    # cleanup again after merge
    combined = combined.dropna(how="all")
    combined = combined[~combined.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)]

    output_file = "cleaned_output.xlsx"
    combined.to_excel(output_file, index=False)

    # formatting
    wb = load_workbook(output_file)
    ws = wb.active

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(output_file)
    print(f"Cleaned file saved: {output_file}")
